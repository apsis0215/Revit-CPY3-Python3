# purpose and IO:
# IN[0]=full file path to XLSX
# IN[1]=worksheet name (string) or None/"" to use first sheet
# IN[2]=full file path to Generic Annotation family template (.rft)
# OUT = dict summary (familyPath, typeCatalogPath, createdParams, createdTypes, warnings)

import os
import re
from collections import defaultdict

import clr

from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import (
    FilteredElementCollector,
    ElementTransformUtils,
    GroupTypeId,
    SpecTypeId,
    XYZ,
    TextNote,
    TextNoteType,
    SaveAsOptions,
    StorageType,
    View
)

doc = DocumentManager.Instance.CurrentDBDocument
uiapp = DocumentManager.Instance.CurrentUIApplication
app = uiapp.Application

xlsx_path = IN[0] if len(IN) > 0 else None
sheet_name = IN[1] if len(IN) > 1 else None
rft_path = IN[2] if len(IN) > 2 else None

warnings = []
result = {
    "familyPath": None,
    "typeCatalogPath": None,
    "createdParams": [],
    "createdTypes": [],
    "warnings": warnings
}

def _is_blank(x):
    return x is None or (isinstance(x, str) and x.strip() == "")

def _is_used_cell(v):
    # Any non-empty value counts (text OR number OR bool)
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True

def _clean_name(name):
    s = "" if name is None else str(name)
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w\-\.\s]", "_", s)
    s = s.strip(" _")
    if len(s) == 0:
        s = "Param"
    return s

def _safe_type_name(name):
    s = _clean_name(name)
    return s if s else "Type"

def _guess_spec_type(col_header, values):
    # Force based on header name (case-insensitive exact match)
    h = "" if col_header is None else str(col_header).strip().lower()
    if h == "yesno":
        return SpecTypeId.Boolean.YesNo
    if h == "number":
        return SpecTypeId.Number
    if h == "integer":
        return SpecTypeId.Int.Integer

    # Otherwise, basic heuristic with Text default
    vals = [v for v in values if v is not None and not (isinstance(v, str) and v.strip() == "")]
    if not vals:
        return SpecTypeId.String.Text

    bool_like = True
    for v in vals:
        if isinstance(v, bool):
            continue
        if isinstance(v, str) and v.strip().lower() in ["true", "false", "yes", "no", "1", "0"]:
            continue
        bool_like = False
        break
    if bool_like:
        return SpecTypeId.Boolean.YesNo

    all_num = True
    all_int = True
    for v in vals:
        if isinstance(v, (int, float)):
            if isinstance(v, float) and abs(v - int(v)) > 1e-9:
                all_int = False
        else:
            try:
                f = float(str(v))
                if abs(f - int(f)) > 1e-9:
                    all_int = False
            except:
                all_num = False
                break

    if all_num:
        return SpecTypeId.Int.Integer if all_int else SpecTypeId.Number

    return SpecTypeId.String.Text

def _to_revit_value(specTypeId, v):
    if v is None:
        return None

    if specTypeId == SpecTypeId.String.Text:
        return str(v)

    if specTypeId == SpecTypeId.Int.Integer:
        try:
            return int(float(v))
        except:
            return None

    if specTypeId == SpecTypeId.Number:
        try:
            return float(v)
        except:
            return None

    if specTypeId == SpecTypeId.Boolean.YesNo:
        if isinstance(v, bool):
            return 1 if v else 0
        s = str(v).strip().lower()
        if s in ["true", "yes", "1"]:
            return 1
        if s in ["false", "no", "0"]:
            return 0
        return None

    return str(v)

def _get_any_view(fdoc):
    try:
        av = fdoc.ActiveView
        if av is not None:
            return av
    except:
        pass

    try:
        views = list(FilteredElementCollector(fdoc).OfClass(View))
        for v in views:
            try:
                if v and not v.IsTemplate:
                    return v
            except:
                continue
    except:
        pass

    return None

def _find_seed_label_and_report(fdoc):
    elems = FilteredElementCollector(fdoc).WhereElementIsNotElementType().ToElements()

    for e in elems:
        try:
            if e.Category and e.Category.Name and e.Category.Name.strip().lower() == "labels":
                return e, ["Found by Category.Name == 'Labels'"]
        except:
            pass

    for e in elems:
        try:
            tn = e.GetType().FullName
            if tn and ("Label" in tn or "FamilyLabel" in tn):
                return e, ["Found by runtime type name: {}".format(tn)]
        except:
            pass

    candidates = []
    for e in elems:
        try:
            for p in e.Parameters:
                try:
                    if p.StorageType == StorageType.ElementId:
                        nm = p.Definition.Name.lower()
                        if "label" in nm and "parameter" in nm:
                            candidates.append((
                                e.Id.IntegerValue,
                                e.GetType().FullName,
                                e.Category.Name if e.Category else None,
                                p.Definition.Name
                            ))
                            break
                except:
                    pass
        except:
            pass

    report = []
    if candidates:
        report.append("Label-like candidates found (ID, Type, Category, ParamName):")
        report.extend([str(x) for x in candidates[:25]])
        report.append("If these are real labels, adjust the finder to return one of them.")
    else:
        report.append("No label-like candidates found. Your template likely has no Label element at all.")
    return None, report

def _try_bind_label(label_elem, family_param):
    try:
        for p in label_elem.Parameters:
            try:
                if p.StorageType == StorageType.ElementId:
                    nm = p.Definition.Name.lower()
                    if "label" in nm and "parameter" in nm:
                        p.Set(family_param.Id)
                        return True
            except:
                pass
    except:
        pass
    return False

# Type catalog token: always OTHER (matches your export)
def _tc_token(_specTypeId):
    return "OTHER"

# VBA-style escaping for CSV
def _vba_csv_escape(val):
    s = "" if val is None else str(val)
    if ("," in s) or ('"' in s) or ("\n" in s) or ("\r" in s):
        s = '"' + s.replace('"', '""') + '"'
    return s

def _ensure_tx(d):
    TransactionManager.Instance.EnsureInTransaction(d)

def _end_tx():
    try:
        TransactionManager.Instance.TransactionTaskDone()
    except:
        pass
    try:
        TransactionManager.Instance.ForceCloseTransaction()
    except:
        pass

def _find_used_bounds_any(ws, max_scan_rows, max_scan_cols):
    last_row = 0
    last_col = 0
    for r in range(1, max_scan_rows + 1):
        row_has_any = False
        for c in range(1, max_scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            if _is_used_cell(v):
                row_has_any = True
                if c > last_col:
                    last_col = c
        if row_has_any and r > last_row:
            last_row = r
    return last_row, last_col

# 1) Validate inputs
if _is_blank(xlsx_path) or not os.path.exists(xlsx_path):
    raise Exception("XLSX path is missing or does not exist: {}".format(xlsx_path))
if _is_blank(rft_path) or not os.path.exists(rft_path):
    raise Exception("Template .rft path is missing or does not exist: {}".format(rft_path))

# 2) Read Excel
try:
    import openpyxl
except Exception as ex:
    raise Exception(
        "openpyxl is required to read XLSX in this Python node.\n"
        "Install/ship openpyxl with Dynamo CPython or use a Zero-Touch node.\nDetails: {}".format(ex)
    )

try:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
except Exception as ex:
    raise Exception("XLSX exists but cannot be opened (locked/corrupt?): {}\n{}".format(xlsx_path, ex))

# 3) Select worksheet
if not _is_blank(sheet_name) and sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    if not _is_blank(sheet_name):
        warnings.append("Worksheet '{}' not found. Using first sheet instead.".format(sheet_name))
    ws = wb[wb.sheetnames[0]]

# 4) Detect bounds by ANY used cell
scan_rows = ws.max_row
scan_cols = ws.max_column
max_row, max_col = _find_used_bounds_any(ws, scan_rows, scan_cols)

if max_row < 2 or max_col < 2:
    raise Exception("Not enough used cells found. Need header row + data + at least 2 columns.")

# 5) Headers row 1
headers_raw = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
headers = [_clean_name(h if h is not None else "Column{}".format(i + 1)) for i, h in enumerate(headers_raw)]

# 6) Data rows 2..max_row
col_values = defaultdict(list)
rows = []
for r in range(2, max_row + 1):
    row_vals = []
    empty_row = True
    for c in range(1, max_col + 1):
        v = ws.cell(row=r, column=c).value
        if not _is_blank(v):
            empty_row = False
        row_vals.append(v)
        col_values[c].append(v)
    if not empty_row:
        rows.append(row_vals)

if not rows:
    raise Exception("No data rows found inside detected bounds.")

# 7) Column X positions in family (feet), min 1/2"
min_gap_in = 0.5
min_gap_ft = min_gap_in / 12.0

widths = []
for c in range(1, max_col + 1):
    letter = openpyxl.utils.get_column_letter(c)
    dim = ws.column_dimensions.get(letter)
    w = None if dim is None else dim.width
    widths.append(float(w) if w is not None else 1.0)

nz = [w for w in widths if w and w > 0]
avg_w = sum(nz) / float(len(nz)) if nz else 1.0
norm = [(w / avg_w) if avg_w > 0 else 1.0 for w in widths]

x_positions_ft = []
x_ft = 0.0
for i in range(len(norm)):
    x_positions_ft.append(x_ft)
    step_ft = max(min_gap_ft, min_gap_ft * norm[i])
    x_ft += step_ft

# 8) Create Family Document
fam_doc = app.NewFamilyDocument(rft_path)

base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
out_dir = os.path.dirname(xlsx_path)
family_path = os.path.join(out_dir, base_name + ".rfa")
typecat_path = os.path.join(out_dir, base_name + ".txt")

# 9) View for TextNote fallback
view = _get_any_view(fam_doc)
if view is None:
    warnings.append("No usable view found in new family document. TextNote fallback will be skipped.")

# 10) Seed label
seed_label_elem, label_report = _find_seed_label_and_report(fam_doc)
seed_label_id = seed_label_elem.Id if seed_label_elem else None
warnings.extend(label_report)
if seed_label_id is None:
    warnings.append(
        "Seed label not found. True label duplication/binding will be skipped. "
        "Add a single Label to the template RFT to enable real labels."
    )

# 11) TextNoteType fallback
text_types = list(FilteredElementCollector(fam_doc).OfClass(TextNoteType))
text_type_id = text_types[0].Id if text_types else None
if text_type_id is None:
    warnings.append("No TextNoteType found in family doc. TextNote fallback will be skipped.")

fm = fam_doc.FamilyManager

# 12) Create TYPE parameters for columns 2..N (Column 1 = type name)
param_map = {}
spec_map = {}

try:
    _ensure_tx(fam_doc)
    for c in range(2, max_col + 1):
        name = headers[c - 1]
        specTypeId = _guess_spec_type(headers[c - 1], col_values[c])
        spec_map[c] = specTypeId

        existing = None
        for p in fm.Parameters:
            if p.Definition.Name == name:
                existing = p
                break
        if existing:
            param_map[c] = existing
            continue

        fp = fm.AddParameter(name, GroupTypeId.Data, specTypeId, False)
        param_map[c] = fp
        result["createdParams"].append(name)
finally:
    _end_tx()

# 13) Place labels/text for columns 2..N
y_ft = 0.0
try:
    _ensure_tx(fam_doc)
    for c in range(2, max_col + 1):
        x = x_positions_ft[c - 1]
        pt = XYZ(x, y_ft, 0)

        if seed_label_id:
            new_ids = ElementTransformUtils.CopyElement(fam_doc, seed_label_id, XYZ(x, 0, 0))
            if new_ids and new_ids.Count > 0:
                new_elem = fam_doc.GetElement(new_ids[0])
                ok = _try_bind_label(new_elem, param_map[c])
                if not ok:
                    warnings.append("Could not bind duplicated label to parameter '{}'.".format(headers[c - 1]))
            else:
                warnings.append("Failed to duplicate seed label for column '{}'.".format(headers[c - 1]))
        else:
            if view is not None and text_type_id is not None:
                try:
                    TextNote.Create(fam_doc, view.Id, pt, headers[c - 1], text_type_id)
                except Exception as ex:
                    warnings.append("Failed to create TextNote for '{}': {}".format(headers[c - 1], ex))
finally:
    _end_tx()

# 14) Create types and set values for columns 2..N
try:
    _ensure_tx(fam_doc)
    for row in rows:
        type_name = _safe_type_name(row[0] if len(row) > 0 else None)

        existing_type = None
        for t in fm.Types:
            if t.Name == type_name:
                existing_type = t
                break

        if existing_type:
            fm.CurrentType = existing_type
        else:
            fm.NewType(type_name)
            result["createdTypes"].append(type_name)

        for c in range(2, max_col + 1):
            fp = param_map.get(c)
            if fp is None:
                continue

            v = row[c - 1] if (c - 1) < len(row) else None
            specTypeId = spec_map.get(c, SpecTypeId.String.Text)
            rv = _to_revit_value(specTypeId, v)
            if rv is None:
                continue

            try:
                if specTypeId == SpecTypeId.String.Text:
                    fm.Set(fp, str(rv))
                elif specTypeId == SpecTypeId.Int.Integer or specTypeId == SpecTypeId.Boolean.YesNo:
                    fm.Set(fp, int(rv))
                elif specTypeId == SpecTypeId.Number:
                    fm.Set(fp, float(rv))
                else:
                    fm.Set(fp, str(rv))
            except Exception as ex:
                warnings.append("Failed to set '{}' for type '{}': {}".format(headers[c - 1], type_name, ex))
finally:
    _end_tx()

# 15) No open phases before SaveAs
try:
    TransactionManager.Instance.ForceCloseTransaction()
except:
    pass

# 16) Save family
opts = SaveAsOptions()
opts.OverwriteExistingFile = True
fam_doc.SaveAs(family_path, opts)
result["familyPath"] = family_path

# 17) Write type catalog like your export:
# Header: leading comma (blank first field), then ParamName##OTHER##
sep = ","

header_fields = [""]
for c in range(2, max_col + 1):
    header_fields.append("{}##{}##".format(headers[c - 1], _tc_token(spec_map.get(c, SpecTypeId.String.Text))))

lines = [sep.join([_vba_csv_escape(x) for x in header_fields])]

for row in rows:
    type_name = _safe_type_name(row[0] if len(row) > 0 else None)
    rec = [type_name]
    for c in range(2, max_col + 1):
        v = row[c - 1] if (c - 1) < len(row) else ""
        if v is None:
            v = ""
        rec.append(v)
    lines.append(sep.join([_vba_csv_escape(x) for x in rec]))

with open(typecat_path, "w", encoding="utf-8-sig") as f:
    f.write("\n".join(lines))

result["typeCatalogPath"] = typecat_path

# 18) Close family doc
fam_doc.Close(False)

OUT = result
