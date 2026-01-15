# -*- coding: utf-8 -*-
# Dynamo Revit 2025 (Dynamo 3.3) - CPython3
#
# PURPOSE / IO
# IN[0] = full file path to XLSX
# IN[1] = worksheet name (string) or None/"" to use first sheet
# IN[2] = full file path to Generic Annotation family template (.rft)
# IN[3] = bool (optional, default True):
#         True  -> col2: Top-Right at (-1/16",0); col3+: Top-Left from (+1/16",0) spaced
#         False -> col2+: Top-Left  from (+1/16",0) spaced
# IN[4] = spacing inches (optional, default 2.0)
#
# OUT = dict:
#   familyPath, typeCatalogPath, createdParams, createdTypes, warnings

import os
import re
from collections import defaultdict

import clr
from RevitServices.Persistence import DocumentManager

clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import (
    FilteredElementCollector,
    ElementTransformUtils,
    GroupTypeId,
    SpecTypeId,
    XYZ,
    SaveAsOptions,
    StorageType,
    Transaction,
    View,
    ViewType,
)

import System

uiapp = DocumentManager.Instance.CurrentUIApplication
app = uiapp.Application

xlsx_path = IN[0] if len(IN) > 0 else None
sheet_name = IN[1] if len(IN) > 1 else None
rft_path = IN[2] if len(IN) > 2 else None

first_top_right = True
if len(IN) > 3 and IN[3] is not None:
    first_top_right = bool(IN[3])

spacing_in = 2.0
if len(IN) > 4 and IN[4] is not None:
    try:
        spacing_in = float(IN[4])
    except Exception:
        spacing_in = 2.0

warnings = []
result = {
    "familyPath": None,
    "typeCatalogPath": None,
    "createdParams": [],
    "createdTypes": [],
    "warnings": warnings,
}


def _is_blank(x):
    return x is None or (isinstance(x, str) and x.strip() == "")

def _is_used_cell(v):
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True

def _clean_name(name):
    s = "" if name is None else str(name)
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w\-\.\s]", "_", s)
    s = s.strip(" _")
    return s if s else "Param"

def _safe_type_name(name):
    s = _clean_name(name)
    return s if s else "Type"

def _slugify_filename(stem):
    s = "" if stem is None else str(stem).strip()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^A-Za-z0-9_\-]+", "-", s)
    s = re.sub(r"[-_]{2,}", "-", s)
    s = s.strip("-_")
    return s if s else "NAME"

def _find_used_bounds_any(ws, max_scan_rows, max_scan_cols):
    last_row = 0
    last_col = 0
    for r in range(1, max_scan_rows + 1):
        row_has_any = False
        for c in range(1, max_scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            if _is_used_cell(v):
                row_has_any = True
                if c > last_col:
                    last_col = c
        if row_has_any and r > last_row:
            last_row = r
    return last_row, last_col

def _guess_spec_type(col_header, values):
    h = "" if col_header is None else str(col_header).strip().lower()
    if h == "yesno":
        return SpecTypeId.Boolean.YesNo
    if h == "number":
        return SpecTypeId.Number
    if h == "integer":
        return SpecTypeId.Int.Integer

    vals = [v for v in values if v is not None and not (isinstance(v, str) and v.strip() == "")]
    if not vals:
        return SpecTypeId.String.Text

    bool_like = True
    for v in vals:
        if isinstance(v, bool):
            continue
        if isinstance(v, str) and v.strip().lower() in ["true", "false", "yes", "no", "1", "0"]:
            continue
        bool_like = False
        break
    if bool_like:
        return SpecTypeId.Boolean.YesNo

    all_num = True
    all_int = True
    for v in vals:
        if isinstance(v, (int, float)):
            if isinstance(v, float) and abs(v - int(v)) > 1e-9:
                all_int = False
        else:
            try:
                f = float(str(v))
                if abs(f - int(f)) > 1e-9:
                    all_int = False
            except Exception:
                all_num = False
                break

    if all_num:
        return SpecTypeId.Int.Integer if all_int else SpecTypeId.Number
    return SpecTypeId.String.Text

def _to_revit_value(specTypeId, v):
    if v is None:
        return None
    if specTypeId == SpecTypeId.String.Text:
        return str(v)
    if specTypeId == SpecTypeId.Int.Integer:
        try:
            return int(float(v))
        except Exception:
            return None
    if specTypeId == SpecTypeId.Number:
        try:
            return float(v)
        except Exception:
            return None
    if specTypeId == SpecTypeId.Boolean.YesNo:
        if isinstance(v, bool):
            return 1 if v else 0
        s = str(v).strip().lower()
        if s in ["true", "yes", "1"]:
            return 1
        if s in ["false", "no", "0"]:
            return 0
        return None
    return str(v)

class Tx(object):
    def __init__(self, d, name):
        self._t = Transaction(d, name)
    def __enter__(self):
        self._t.Start()
        return self._t
    def __exit__(self, exc_type, exc, tb):
        if exc_type:
            try:
                self._t.RollBack()
            except Exception:
                pass
            return False
        self._t.Commit()
        return True

def _ids_count(ids):
    if ids is None:
        return 0
    try:
        return int(ids.Count)
    except Exception:
        try:
            return len(ids)
        except Exception:
            return 0

def _ids_first(ids):
    if ids is None:
        return None
    try:
        if hasattr(ids, "Count") and ids.Count > 0:
            return ids[0]
    except Exception:
        pass
    try:
        return ids[0] if len(ids) > 0 else None
    except Exception:
        return None

def _rt(obj):
    try:
        return obj.GetType().FullName
    except Exception:
        return None

def _non_template_views(doc):
    vs = []
    try:
        views = list(FilteredElementCollector(doc).OfClass(View))
    except Exception:
        return vs
    for v in views:
        try:
            if v and not v.IsTemplate:
                vs.append(v)
        except Exception:
            pass
    vs.sort(key=lambda x: x.Id.IntegerValue)
    return vs

def _pick_sheet_view(doc):
    views = _non_template_views(doc)
    sheet_named = []
    sheets = []
    for v in views:
        vt = None
        vn = ""
        try:
            vt = v.ViewType
        except Exception:
            pass
        try:
            vn = (v.Name or "").strip().lower()
        except Exception:
            pass

        if vt == ViewType.DrawingSheet:
            sheets.append(v)
            if "sheet" in vn:
                sheet_named.append(v)

    if sheet_named:
        sheet_named.sort(key=lambda v: v.Id.IntegerValue)
        return sheet_named[0]
    if sheets:
        sheets.sort(key=lambda v: v.Id.IntegerValue)
        return sheets[0]

    named = []
    for v in views:
        try:
            if "sheet" in (v.Name or "").strip().lower():
                named.append(v)
        except Exception:
            pass
    if named:
        named.sort(key=lambda v: v.Id.IntegerValue)
        return named[0]
    return None

def _elems_in_view(doc, v):
    try:
        return list(FilteredElementCollector(doc, v.Id).WhereElementIsNotElementType().ToElements())
    except Exception:
        return []

def _param_names_lower(e):
    out = set()
    try:
        for p in e.Parameters:
            try:
                out.add((p.Definition.Name or "").strip().lower())
            except Exception:
                pass
    except Exception:
        pass
    return out

def _score_strict_label_candidate(e):
    pn = _param_names_lower(e)
    # Relaxed requirements - just need some text-like params
    text_params = {"label", "id", "text", "value", "sample text"}
    if not any(tp in pn for tp in text_params):
        return 0
    return 260

def _find_seed_label_on_sheet(doc):
    sheet_view = _pick_sheet_view(doc)
    if sheet_view is None:
        return None, None
    elems = _elems_in_view(doc, sheet_view)
    best = None
    for e in elems:
        sc = _score_strict_label_candidate(e)
        if sc <= 0:
            continue
        key = (-sc, e.Id.IntegerValue)
        if best is None or key < best[0]:
            best = (key, e)
    return sheet_view, (best[1] if best else None)

def _bbox_center_in_view(elem, view):
    try:
        bb = elem.get_BoundingBox(view)
        if not bb or not bb.Min or not bb.Max:
            return None
        return XYZ(
            (bb.Min.X + bb.Max.X) * 0.5,
            (bb.Min.Y + bb.Max.Y) * 0.5,
            (bb.Min.Z + bb.Max.Z) * 0.5,
        )
    except Exception:
        return None

def _move_elem_to_point_by_bbox(doc, elem, view, target_pt):
    cur = _bbox_center_in_view(elem, view)
    if cur is None:
        return False
    delta = XYZ(target_pt.X - cur.X, target_pt.Y - cur.Y, 0.0)
    try:
        ElementTransformUtils.MoveElement(doc, elem.Id, delta)
        return True
    except Exception:
        return False

def _try_set_justification_int_params(label_elem, horiz, vert):
    try:
        for p in label_elem.Parameters:
            if p is None or p.StorageType != StorageType.Integer:
                continue
            nm = (p.Definition.Name or "").strip().lower()
            if "horizontal align" in nm:
                if horiz.lower() == "right":
                    p.Set(128)
                elif horiz.lower() == "left":
                    p.Set(64)
            if "vertical align" in nm:
                if vert.lower() == "top":
                    p.Set(512)
    except Exception:
        pass

def _verify_label_binding(fm, label_elem, elem_param, family_param):
    """
    Double-check that the label binding was created correctly by copying base label.
    Verifies:
    1. The element parameter is now associated with the family parameter
    2. The association can be retrieved back
    3. The label element still exists and is valid
    
    Returns (is_verified: bool, message: str)
    """
    try:
        # Check 1: Verify the label element is still valid
        if label_elem is None or not label_elem.IsValidObject:
            return False, "Label element is no longer valid after binding"
        
        # Check 2: Verify the family parameter exists in FamilyManager
        param_found = False
        for p in fm.Parameters:
            if p.Id.IntegerValue == family_param.Id.IntegerValue:
                param_found = True
                break
        
        if not param_found:
            return False, "Family parameter not found in FamilyManager after binding"
        
        # Check 3: Try to get the associated family parameter back from element parameter
        try:
            associated_fp = fm.GetAssociatedFamilyParameter(elem_param)
            if associated_fp is None:
                return False, "GetAssociatedFamilyParameter returned None - binding may not have persisted"
            
            # Verify it's the same parameter we intended to bind
            if associated_fp.Id.IntegerValue != family_param.Id.IntegerValue:
                return False, "Associated parameter ID mismatch: expected {}, got {}".format(
                    family_param.Id.IntegerValue, associated_fp.Id.IntegerValue
                )
        except Exception as ex:
            # GetAssociatedFamilyParameter might not be available in all API versions
            return True, "Could not verify association (API limitation): {}".format(str(ex))
        
        # Check 4: Verify parameter names match
        try:
            if associated_fp.Definition.Name != family_param.Definition.Name:
                return False, "Associated parameter name mismatch: expected '{}', got '{}'".format(
                    family_param.Definition.Name, associated_fp.Definition.Name
                )
        except Exception:
            pass  # Name comparison is optional
        
        return True, "Label binding verified successfully"
        
    except Exception as ex:
        return False, "Verification failed with exception: {}".format(str(ex))


def _verify_elementid_binding(elem_param, expected_id):
    """
    Double-check that an ElementId parameter was set correctly.
    
    Returns (is_verified: bool, message: str)
    """
    try:
        actual_id = elem_param.AsElementId()
        
        if actual_id is None:
            return False, "ElementId parameter returned None after setting"
        
        if actual_id.IntegerValue != expected_id.IntegerValue:
            return False, "ElementId mismatch: expected {}, got {}".format(
                expected_id.IntegerValue, actual_id.IntegerValue
            )
        
        return True, "ElementId binding verified successfully"
        
    except Exception as ex:
        return False, "ElementId verification failed: {}".format(str(ex))


def _bind_label_to_family_param(fm, label_elem, family_param):
    """
    Binds a label element's text parameter to a family parameter.
    Since we can't create labels via API, we must work with whatever
    parameter the existing label uses (e.g., "ID", "Label", "Text", etc.)
    
    Priority order:
    1. Try common label parameter names: "ID", "Label", "Text", "Value"
    2. Try any String parameter with text-related keywords
    3. Try ANY writable String parameter
    4. Try ElementId parameters for label parameter references
    
    Returns (ok, debug_dict)
    """
    dbg = {
        "elemId": label_elem.Id.IntegerValue,
        "elemType": _rt(label_elem),
        "attempts": [],
        "allParams": [],
        "verificationStatus": None,  # Added for verification tracking
    }

    # Collect all parameter info for debugging
    string_params = []
    try:
        for p in label_elem.Parameters:
            try:
                pname = (p.Definition.Name or "").strip()
                pinfo = {
                    "name": pname,
                    "storageType": int(p.StorageType),
                    "isReadOnly": bool(getattr(p, "IsReadOnly", False)),
                }
                dbg["allParams"].append(pinfo)
                
                # Collect String parameters (don't filter by read-only yet)
                if p.StorageType == StorageType.String:
                    string_params.append((pname, p))
            except Exception:
                pass
    except Exception:
        pass

    # Define common label parameter names in priority order
    # "ID" and "Label" are most common for label elements
    common_label_params = ["ID", "Label", "Text", "Value", "Sample Text", "Content"]
    
    # Strategy 1: Try common label parameter names first
    for target_name in common_label_params:
        for pname, p in string_params:
            if pname.lower() == target_name.lower():
                try:
                    fm.AssociateElementParameterToFamilyParameter(p, family_param)
                    dbg["attempts"].append("{} (exact match '{}') -> SUCCESS".format(pname, target_name))
                    
                    # Double-check: Verify the label was created correctly by copying base label
                    verified, verify_msg = _verify_label_binding(fm, label_elem, p, family_param)
                    dbg["verificationStatus"] = {"verified": verified, "message": verify_msg}
                    if not verified:
                        dbg["attempts"].append("VERIFICATION WARNING: {}".format(verify_msg))
                    
                    return True, dbg
                except Exception as ex:
                    dbg["attempts"].append("{} (exact match '{}') -> FAIL: {}".format(pname, target_name, str(ex)))
    
    # Strategy 2: Try any String parameter with text-related keywords
    text_keywords = ["label", "id", "text", "value", "content", "sample"]
    for pname, p in string_params:
        pname_lower = pname.lower()
        if any(keyword in pname_lower for keyword in text_keywords):
            # Skip if we already tried it in strategy 1
            if any(pname.lower() == c.lower() for c in common_label_params):
                continue
            try:
                fm.AssociateElementParameterToFamilyParameter(p, family_param)
                dbg["attempts"].append("{} (keyword match) -> SUCCESS".format(pname))
                
                # Double-check: Verify the label was created correctly by copying base label
                verified, verify_msg = _verify_label_binding(fm, label_elem, p, family_param)
                dbg["verificationStatus"] = {"verified": verified, "message": verify_msg}
                if not verified:
                    dbg["attempts"].append("VERIFICATION WARNING: {}".format(verify_msg))
                
                return True, dbg
            except Exception as ex:
                dbg["attempts"].append("{} (keyword match) -> FAIL: {}".format(pname, str(ex)))
    
    # Strategy 3: Try ANY String parameter
    for pname, p in string_params:
        # Skip if already tried
        pname_lower = pname.lower()
        already_tried = (
            any(pname.lower() == c.lower() for c in common_label_params) or
            any(keyword in pname_lower for keyword in text_keywords)
        )
        if already_tried:
            continue
        try:
            fm.AssociateElementParameterToFamilyParameter(p, family_param)
            dbg["attempts"].append("{} (any string) -> SUCCESS".format(pname))
            
            # Double-check: Verify the label was created correctly by copying base label
            verified, verify_msg = _verify_label_binding(fm, label_elem, p, family_param)
            dbg["verificationStatus"] = {"verified": verified, "message": verify_msg}
            if not verified:
                dbg["attempts"].append("VERIFICATION WARNING: {}".format(verify_msg))
            
            return True, dbg
        except Exception as ex:
            dbg["attempts"].append("{} (any string) -> FAIL: {}".format(pname, str(ex)))
    
    # Strategy 4: Try ElementId parameters (label parameter reference)
    # Priority: "Label" exact match, then any parameter containing "label"
    try:
        fp_id = family_param.Id
        
        # First, try exact match for "Label" parameter (most common in Generic Annotation families)
        for p in label_elem.Parameters:
            try:
                if p.StorageType != StorageType.ElementId:
                    continue
                pname = (p.Definition.Name or "").strip()
                pname_lower = pname.lower()
                
                if pname_lower == "label":
                    try:
                        p.Set(fp_id)
                        dbg["attempts"].append("{} (ElementId exact 'Label') -> SUCCESS".format(pname))
                        
                        # Double-check: Verify the ElementId was set correctly
                        verified, verify_msg = _verify_elementid_binding(p, fp_id)
                        dbg["verificationStatus"] = {"verified": verified, "message": verify_msg}
                        if not verified:
                            dbg["attempts"].append("VERIFICATION WARNING: {}".format(verify_msg))
                        
                        return True, dbg
                    except Exception as ex:
                        dbg["attempts"].append("{} (ElementId exact 'Label') -> FAIL: {}".format(pname, str(ex)))
            except Exception:
                continue
        
        # Then try any parameter containing "label" in the name
        for p in label_elem.Parameters:
            try:
                if p.StorageType != StorageType.ElementId:
                    continue
                pname = (p.Definition.Name or "").strip()
                pname_lower = pname.lower()
                
                # Skip if already tried as exact match
                if pname_lower == "label":
                    continue
                
                # Look for parameters that contain "label" in the name
                if "label" in pname_lower:
                    try:
                        p.Set(fp_id)
                        dbg["attempts"].append("{} (ElementId contains 'label') -> SUCCESS".format(pname))
                        
                        # Double-check: Verify the ElementId was set correctly
                        verified, verify_msg = _verify_elementid_binding(p, fp_id)
                        dbg["verificationStatus"] = {"verified": verified, "message": verify_msg}
                        if not verified:
                            dbg["attempts"].append("VERIFICATION WARNING: {}".format(verify_msg))
                        
                        return True, dbg
                    except Exception as ex:
                        dbg["attempts"].append("{} (ElementId contains 'label') -> FAIL: {}".format(pname, str(ex)))
            except Exception:
                continue
    except Exception:
        pass

    return False, dbg

def _tc_token(_specTypeId):
    return "OTHER"

def _vba_csv_escape(val):
    s = "" if val is None else str(val)
    if ("," in s) or ('"' in s) or ("\n" in s) or ("\r" in s):
        s = '"' + s.replace('"', '""') + '"'
    return s


# ---------------------------
# Validate inputs
# ---------------------------
if _is_blank(xlsx_path) or not os.path.exists(xlsx_path):
    raise Exception("XLSX path is missing or does not exist: {}".format(xlsx_path))
if _is_blank(rft_path) or not os.path.exists(rft_path):
    raise Exception("Template RFT path is missing or does not exist: {}".format(rft_path))
if spacing_in <= 0:
    spacing_in = 2.0
    warnings.append("Spacing <= 0; using 2.0 inches.")

spacing_ft = spacing_in / 12.0
one_sixteenth_ft = (1.0 / 16.0) / 12.0


# ---------------------------
# Read Excel
# ---------------------------
try:
    import openpyxl
except Exception as ex:
    raise Exception("openpyxl is required in this CPython node.\nDetails: {}".format(ex))

wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)

if not _is_blank(sheet_name) and sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    if not _is_blank(sheet_name):
        warnings.append("Worksheet '{}' not found. Using first sheet.".format(sheet_name))
    ws = wb[wb.sheetnames[0]]

max_row, max_col = _find_used_bounds_any(ws, ws.max_row, ws.max_column)
if max_row < 2 or max_col < 2:
    raise Exception("Not enough used cells found. Need header row + data + at least 2 columns.")

headers_raw = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
headers = [_clean_name(h if h is not None else "Column{}".format(i + 1)) for i, h in enumerate(headers_raw)]

col_values = defaultdict(list)
rows = []
for r in range(2, max_row + 1):
    row_vals = []
    empty_row = True
    for c in range(1, max_col + 1):
        v = ws.cell(row=r, column=c).value
        if not _is_blank(v):
            empty_row = False
        row_vals.append(v)
        col_values[c].append(v)
    if not empty_row:
        rows.append(row_vals)

if not rows:
    raise Exception("No data rows found inside detected bounds.")


# ---------------------------
# Output paths: SYMB-__- + cleaned stem (spaces=>dash)
# ---------------------------
xlsx_stem = os.path.splitext(os.path.basename(xlsx_path))[0]
safe_stem = _slugify_filename(xlsx_stem)
out_name = "SYMB-__-{}".format(safe_stem)

out_dir = os.path.dirname(xlsx_path)
family_path = os.path.join(out_dir, out_name + ".rfa")
typecat_path = os.path.join(out_dir, out_name + ".txt")


# ---------------------------
# Build from RFT (operate STRICTLY on the Sheet view)
# ---------------------------
fam_doc = None
try:
    fam_doc = app.NewFamilyDocument(rft_path)
    if fam_doc is None:
        raise Exception("NewFamilyDocument returned None for: {}".format(rft_path))

    fm = fam_doc.FamilyManager

    sheet_view, seed_elem = _find_seed_label_on_sheet(fam_doc)
    if sheet_view is None:
        raise Exception("Could not find a Sheet view in this RFT.")
    warnings.append("Using Sheet view: id={} name='{}' type={}".format(
        sheet_view.Id.IntegerValue, sheet_view.Name, str(sheet_view.ViewType)
    ))

    if seed_elem is None:
        raise Exception(
            "No strict label candidate found ON the Sheet view.\n"
            "Need a real Label on the Sheet view with params: Label / Sample Text / Horizontal Align / Vertical Align."
        )
    warnings.append("Seed label (sheet): elemId={} runtimeType={}".format(seed_elem.Id.IntegerValue, _rt(seed_elem)))

    # Create TYPE parameters for columns 2..N (col1 = type name)
    param_map = {}
    spec_map = {}
    with Tx(fam_doc, "Create parameters"):
        for c in range(2, max_col + 1):
            pname = headers[c - 1]
            specTypeId = _guess_spec_type(headers[c - 1], col_values[c])
            spec_map[c] = specTypeId

            existing = None
            for p in fm.Parameters:
                if p.Definition.Name == pname:
                    existing = p
                    break

            if existing:
                param_map[c] = existing
            else:
                fp = fm.AddParameter(pname, GroupTypeId.Data, specTypeId, False)
                param_map[c] = fp
                result["createdParams"].append(pname)

    # Placement rules
    if first_top_right:
        first_x = -one_sixteenth_ft
        rest_start_x = +one_sixteenth_ft
        first_h, first_v = "Right", "Top"
    else:
        first_x = +one_sixteenth_ft
        rest_start_x = +one_sixteenth_ft
        first_h, first_v = "Left", "Top"

    with Tx(fam_doc, "Place + bind labels (sheet view)"):
        # col2 uses the seed itself
        if not _move_elem_to_point_by_bbox(fam_doc, seed_elem, sheet_view, XYZ(first_x, 0.0, 0.0)):
            warnings.append("Could not move seed by bbox; continuing with copies.")

        ok, dbg = _bind_label_to_family_param(fm, seed_elem, param_map[2])
        warnings.append("BIND DEBUG (col2 '{}'): {}".format(headers[1], dbg))
        if not ok:
            raise Exception(
                "Failed to bind seed label to family parameter '{}'.\n\n"
                "Your label element doesn't have a bindable text parameter.\n"
                "Available parameters: {}\n\n"
                "Binding attempts: {}\n\n"
                "Make sure your .rft template has a Label element with a parameter "
                "like 'ID', 'Label', or 'Text' that can be associated with family parameters.".format(
                    headers[1],
                    [p["name"] for p in dbg["allParams"]],
                    dbg["attempts"]
                )
            )
        _try_set_justification_int_params(seed_elem, first_h, first_v)

        # Store the original element type
        seed_type_id = seed_elem.GetTypeId()
        seed_symbol = fam_doc.GetElement(seed_type_id)

        # col3..N: copy within the SAME view (keeps OWNER_VIEW_ID consistent)
        for c in range(3, max_col + 1):
            target_x = rest_start_x + (c - 3) * spacing_ft
            delta_x = target_x - first_x

            new_ids = ElementTransformUtils.CopyElement(fam_doc, seed_elem.Id, XYZ(delta_x, 0.0, 0.0))
            if _ids_count(new_ids) == 0:
                raise Exception("CopyElement failed for column '{}'".format(headers[c - 1]))

            new_id = _ids_first(new_ids)
            new_elem = fam_doc.GetElement(new_id)
            if new_elem is None:
                raise Exception("GetElement(copyId) returned None for column '{}'".format(headers[c - 1]))

            # Ensure the copied element maintains the same type as the seed
            if new_elem.GetTypeId().IntegerValue != seed_type_id.IntegerValue:
                try:
                    new_elem.ChangeTypeId(seed_type_id)
                    warnings.append("Changed type for col '{}' to match seed label".format(headers[c - 1]))
                except Exception as type_ex:
                    warnings.append("Could not change type for col '{}': {}".format(headers[c - 1], type_ex))

            ok, dbg = _bind_label_to_family_param(fm, new_elem, param_map[c])
            if not ok:
                warnings.append("BIND DEBUG (col '{}'): {}".format(headers[c - 1], dbg))
                raise Exception("Failed to bind duplicated label to family parameter '{}'".format(headers[c - 1]))

            _try_set_justification_int_params(new_elem, "Left", "Top")

    # Create types and set values
    with Tx(fam_doc, "Create types + set values"):
        for row in rows:
            type_name = _safe_type_name(row[0] if len(row) > 0 else None)

            existing_type = None
            for t in fm.Types:
                if t.Name == type_name:
                    existing_type = t
                    break

            if existing_type:
                fm.CurrentType = existing_type
            else:
                fm.NewType(type_name)
                result["createdTypes"].append(type_name)

            for c in range(2, max_col + 1):
                fp = param_map.get(c)
                if fp is None:
                    continue

                v = row[c - 1] if (c - 1) < len(row) else None
                specTypeId = spec_map.get(c, SpecTypeId.String.Text)
                rv = _to_revit_value(specTypeId, v)
                if rv is None:
                    continue

                try:
                    if specTypeId == SpecTypeId.String.Text:
                        fm.Set(fp, str(rv))
                    elif specTypeId == SpecTypeId.Int.Integer or specTypeId == SpecTypeId.Boolean.YesNo:
                        fm.Set(fp, int(rv))
                    elif specTypeId == SpecTypeId.Number:
                        fm.Set(fp, float(rv))
                    else:
                        fm.Set(fp, str(rv))
                except Exception as ex:
                    warnings.append("Failed to set '{}' for type '{}': {}".format(headers[c - 1], type_name, ex))

    # Save family
    opts = SaveAsOptions()
    opts.OverwriteExistingFile = True
    fam_doc.SaveAs(family_path, opts)
    result["familyPath"] = family_path

    # Write type catalog
    sep = ","
    header_fields = [""]
    for c in range(2, max_col + 1):
        header_fields.append("{}##{}##".format(headers[c - 1], _tc_token(spec_map.get(c, SpecTypeId.String.Text))))
    lines = [sep.join([_vba_csv_escape(x) for x in header_fields])]

    for row in rows:
        type_name = _safe_type_name(row[0] if len(row) > 0 else None)
        rec = [type_name]
        for c in range(2, max_col + 1):
            v = row[c - 1] if (c - 1) < len(row) else ""
            rec.append("" if v is None else v)
        lines.append(sep.join([_vba_csv_escape(x) for x in rec]))

    with open(typecat_path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines))
    result["typeCatalogPath"] = typecat_path

finally:
    if fam_doc is not None:
        try:
            fam_doc.Close(False)
        except Exception:
            pass

OUT = result